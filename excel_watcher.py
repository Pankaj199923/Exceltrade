"""
Excel Watcher
-----------------------------------
Ye script tumhare laptop par background me chalti rahegi.
Jaise hi tum Excel file save karoge, ye automatically:
  1. Latest data padhega (openpyxl se)
  2. Server ko POST kar dega (/update endpoint)

Chalane se pehle neeche CONFIG section me apni values daal do.
"""

import time
import json
from datetime import datetime

import requests
from openpyxl import load_workbook
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ============ CONFIG — YAHAN APNI VALUES DAALO ============
EXCEL_FILE_PATH = r"Z:\PANKAJ\pankaj_codes\Strategy\opch.xlsx"  # apni Excel file ka pura path
SHEET_NAME = None          # None = active/first sheet, ya "Sheet1" jaisa naam do
SERVER_URL = "http://localhost:5000/update"   # deploy karne ke baad live URL yahan daalo
API_KEY = "mysecretkey123"                 # server.py wali API_KEY se match honi chahiye
DEBOUNCE_SECONDS = 2        # itni der wait karega taaki Excel poora save ho jaaye
# ============================================================


def read_excel_data():
    """Excel file se headers + rows nikalta hai (JSON-safe format me)."""
    wb = load_workbook(EXCEL_FILE_PATH, data_only=True)  # data_only=True => formula ka result milega
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = list(next(rows_iter))  # pehli row = headers

    data_rows = []
    for row in rows_iter:
        if all(cell is None for cell in row):
            continue  # khali row skip
        data_rows.append(["" if c is None else c for c in row])

    return {
        "sheet_name": ws.title,
        "headers": ["" if h is None else str(h) for h in headers],
        "rows": data_rows,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def send_to_server(data):
    try:
        resp = requests.post(
            SERVER_URL,
            headers={"X-API-Key": API_KEY, "Content-Type": "application/json"},
            data=json.dumps(data, default=str),
            timeout=10,
        )
        if resp.status_code == 200:
            print(f"[OK] Sent {len(data['rows'])} rows to server @ {data['updated_at']}")
        else:
            print(f"[ERROR] Server responded {resp.status_code}: {resp.text}")
    except requests.exceptions.RequestException as e:
        print(f"[ERROR] Could not reach server: {e}")


class ExcelChangeHandler(FileSystemEventHandler):
    def __init__(self):
        self.last_sent = 0

    def on_modified(self, event):
        if event.is_directory:
            return
        if not event.src_path.replace("/", "\\").endswith(EXCEL_FILE_PATH.split("\\")[-1]):
            return

        now = time.time()
        if now - self.last_sent < DEBOUNCE_SECONDS:
            return  # bahut jaldi jaldi trigger na ho (Excel save 2-3 baar event fire karta hai)
        self.last_sent = now

        # Excel poora likh chuke, isliye thoda ruk ke padho
        time.sleep(0.5)
        try:
            data = read_excel_data()
            send_to_server(data)
        except Exception as e:
            print(f"[ERROR] Failed to read/send Excel data: {e}")


def main():
    import os
    folder = os.path.dirname(EXCEL_FILE_PATH) or "."

    print(f"Watching: {EXCEL_FILE_PATH}")
    print(f"Sending updates to: {SERVER_URL}")

    # Start hote hi ek baar turant bhej do
    try:
        send_to_server(read_excel_data())
    except Exception as e:
        print(f"[WARN] Initial read failed: {e}")

    event_handler = ExcelChangeHandler()
    observer = Observer()
    observer.schedule(event_handler, folder, recursive=False)
    observer.start()
    print("Watcher chalu hai... (band karne ke liye Ctrl+C dabao)")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


if __name__ == "__main__":
    main()
